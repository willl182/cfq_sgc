import openpyxl
import os
import pandas as pd

def inspect_xlsx(filepath):
    print(f"\n=================== INSPECTING: {os.path.basename(filepath)} ===================")
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        sheets = wb.sheetnames
        print(f"Sheets: {sheets}")
        for sheet in sheets[:3]:  # inspect first 3 sheets
            ws = wb[sheet]
            print(f"Sheet '{sheet}': {ws.max_row} rows x {ws.max_column} cols")
            # print first 5 rows
            rows = list(ws.iter_rows(max_row=5, values_only=True))
            for r in rows:
                print(f"  {r}")
    except Exception as e:
        print(f"Error inspecting {filepath}: {e}")

db_dir = "/home/w182/w421/cfq_sgc/SGC_CALFERQUIM/08_Base_Datos_Tecnica"
files = [
    "Inventario_Materias_Primas_Dossier_V1.csv",
    "Composicion_Ingredientes_ICA.xlsx",
    "F-001-FERTILIZANTES_DILIGENCIADO_20260429.xlsx"
]

for f in files:
    path = os.path.join(db_dir, f)
    if os.path.exists(path):
        if path.endswith('.csv'):
            print(f"\n=================== INSPECTING: {f} ===================")
            df = pd.read_csv(path)
            print(df.head())
        else:
            inspect_xlsx(path)
    else:
        print(f"File not found: {path}")
